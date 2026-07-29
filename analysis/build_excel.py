"""
Build multi-sheet Excel workbook for Activation-Steered-Personas experiment comparison.

Experiments
-----------
Exp1  baseline          : user=local,    system=local
Exp2  sa      : user=local,    system=sa
Exp3  user-steer        : user=steered,  system=local        (40 runs: 10 traits × 4 scalars)
Exp4  user-steer+sa     : user=steered,  system=sa (40 runs)

Scalars:  s-1 (anti), s0 (neutral≡no-steer), s1 (mild), s2 (strong)
Traits:   calm, careless, compassionate, consistent, dependable,
          inventive, nervous, outgoing, self-interested, solitary

Usage
-----
  # Default (legacy qwen paths):
  python analysis/build_excel.py

  # Custom experiment root:
  python analysis/build_excel.py \\
    --exp-root /work/hdd/beto/balvesrodrigues/experiments-persona/llama-v2 \\
    --baseline     llama-baseline-2026-05-27-07-42 \\
    --sa llama-sa-2026-05-27-07-52 \\
    --user-steer   llama-user-steer-2026-05-27-07-43 \\
    --user-steer-sa llama-user-steer-sa-both-2026-05-27-07-43 \\
    --output /work/hdd/beto/balvesrodrigues/experiments-persona/llama-v2/analysis/llama_v2_results.xlsx \\
    --title "LLaMA-3.1-8B"
"""

import argparse
import json
import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ── defaults (legacy qwen paths) ──────────────────────────────────────────────

_DEFAULT_BASE = Path("/work/hdd/beto/balvesrodrigues/persona_vectors/Activation-Steered-Personas/output")
_DEFAULT_EXPERIMENTS = {
    "baseline":      "baseline-2026-05-13-15-11",
    "sa":  "sa-2026-05-13-17-24",
    "user-steer":    "user-steer-2026-05-13-16-26",
    "user-steer-sa": "user-steer-sa-both-2026-05-13-16-26",
}
_DEFAULT_OUTPUT = str(Path(__file__).parent / "persona_tod_results.xlsx")

TRAITS  = ["calm", "careless", "compassionate", "consistent", "dependable",
           "inventive", "nervous", "outgoing", "self-interested", "solitary"]
SCALARS = ["s-1", "s0", "s1", "s2"]

# ── metric definitions ─────────────────────────────────────────────────────────
# (key, label, display_mode)
#   display_mode: "pct"  → stored 0-1, shown as %
#                 "pct100" → stored 0-100, shown as %  (divide by 100)
#                 "raw"  → raw float, shown as 0.000
#                 "ppl"  → perplexity, shown as 0.00 (lower=better)

METRICS = [
    ("method_accuracy",              "Method Acc",       "pct"),
    ("key_accuracy",                 "Key Acc",          "pct"),
    ("key_iou",                      "Key IoU",          "pct"),
    ("value_accuracy",               "Value Acc",        "pct"),
    ("value_iou",                    "Value IoU",        "pct"),
    ("full_api_accuracy",            "Full API Acc",     "pct"),
    ("full_api_accuracy_iou",        "Full API IoU",     "pct"),
    ("successful_dialogs_rate",      "Dialog Success",   "pct"),
    ("successful_dialogs_iou_rate",  "Dialog Succ IoU",  "pct"),
    ("inform_accuracy",              "Inform Acc",       "pct"),
    ("inform_accuracy_with_search_results", "Inform+Search", "pct"),
    ("bleu",                         "BLEU",             "pct"),
    ("dialog_completion_rate",       "Completion",       "pct"),
    ("bert_score.avg_f1",            "BERT F1",          "pct"),
    ("bert_score_user.avg_f1",       "BERT User F1",     "pct"),
    ("bert_score_system.avg_f1",     "BERT System F1",   "pct"),
    # LLM-judged quality
    ("constraint_satisfaction_pct",  "Csat",             "pct100"),
    ("inform_score",                 "Inform (LLM)",     "raw"),
    ("truthfulness_score",           "Truthful",         "raw"),
    ("user_satisfaction_score",      "User Sat",         "raw"),
    # Perplexity
    ("perplexity.perplexity",        "PPL",              "ppl"),
]

METRIC_KEYS   = [m[0]  for m in METRICS]
METRIC_LABELS = [m[1]  for m in METRICS]
METRIC_MODES  = {m[0]: m[2] for m in METRICS}

# ── style constants ────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER2_FILL = PatternFill("solid", fgColor="2E75B6")
SUBHEAD_FILL = PatternFill("solid", fgColor="D6E4F0")
EXP_FILLS = {
    "baseline":      PatternFill("solid", fgColor="E2EFDA"),
    "sa":  PatternFill("solid", fgColor="FFF2CC"),
    "user-steer":    PatternFill("solid", fgColor="DDEBF7"),
    "user-steer-sa": PatternFill("solid", fgColor="FCE4D6"),
}
SCALAR_FILLS = {
    "s-1": PatternFill("solid", fgColor="FADBD8"),
    "s0":  PatternFill("solid", fgColor="F5F5F5"),
    "s1":  PatternFill("solid", fgColor="D5F5E3"),
    "s2":  PatternFill("solid", fgColor="D6EAF8"),
}
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
LABEL_FONT  = Font(bold=True, size=10)
NORMAL_FONT = Font(size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center")


def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def style_header(cell, fill=None, font=None):
    cell.fill      = fill or HEADER_FILL
    cell.font      = font or HEADER_FONT
    cell.alignment = CENTER
    cell.border    = thin_border()


def style_data(cell, fill=None):
    cell.fill      = fill or PatternFill()
    cell.font      = NORMAL_FONT
    cell.alignment = CENTER
    cell.border    = thin_border()


def _display_value(key, raw_val):
    """Convert raw metric value to displayable (cell value, number_format)."""
    if raw_val is None:
        return None, "General"
    mode = METRIC_MODES.get(key, "pct")
    if mode == "pct":
        return float(raw_val), "0.00%"
    elif mode == "pct100":
        return float(raw_val) / 100.0, "0.00%"
    elif mode == "raw":
        return float(raw_val), "0.000"
    elif mode == "ppl":
        return float(raw_val), "0.00"
    return raw_val, "General"


def write_metric(ws, row, col, key, raw_val, fill=None):
    val, fmt = _display_value(key, raw_val)
    cell = ws.cell(row=row, column=col, value=val)
    style_data(cell, fill)
    cell.number_format = fmt


def write_delta(ws, row, col, key, delta):
    if delta is None:
        cell = ws.cell(row=row, column=col, value="")
        cell.alignment = CENTER
        cell.border = thin_border()
        return
    mode = METRIC_MODES.get(key, "pct")
    if mode in ("pct", "pct100"):
        if mode == "pct100":
            delta = delta / 100.0
        fmt = "+0.00%;-0.00%;0.00%"
        cell_val = delta
    else:
        fmt = "+0.000;-0.000;0.000"
        cell_val = delta
    cell = ws.cell(row=row, column=col, value=cell_val)
    cell.font      = Font(size=10, bold=True,
                          color="375623" if delta >= 0 else "C00000")
    cell.alignment = CENTER
    cell.border    = thin_border()
    cell.number_format = fmt


def apply_color_scale(ws, first_row, last_row, col):
    col_letter = get_column_letter(col)
    ref  = f"{col_letter}{first_row}:{col_letter}{last_row}"
    rule = ColorScaleRule(
        start_type="percentile", start_value=10,  start_color="F8696B",
        mid_type="percentile",   mid_value=50,    mid_color="FFEB84",
        end_type="percentile",   end_value=90,    end_color="63BE7B",
    )
    ws.conditional_formatting.add(ref, rule)


def freeze_and_zoom(ws, cell="B2", zoom=100):
    ws.freeze_panes    = cell
    ws.sheet_view.zoomScale = zoom


# ── data loading ───────────────────────────────────────────────────────────────

def load_metrics(path: Path) -> dict:
    with open(path / "metrics.json") as f:
        d = json.load(f)["summary"]
    row = {}
    for key in METRIC_KEYS:
        if "." in key:
            a, b = key.split(".", 1)
            row[key] = d.get(a, {}).get(b) if isinstance(d.get(a), dict) else None
        else:
            row[key] = d.get(key)
    return row


def load_all(experiments: dict) -> pd.DataFrame:
    rows = []

    m = load_metrics(experiments["baseline"])
    rows.append({"exp": "baseline", "trait": "—", "scalar": "—", "run": "baseline", **m})

    m = load_metrics(experiments["sa"])
    rows.append({"exp": "sa", "trait": "—", "scalar": "—", "run": "sa", **m})

    for exp_key in ("user-steer", "user-steer-sa"):
        for trait in TRAITS:
            for scalar in SCALARS:
                run_name = f"{trait}__{scalar}"
                p = experiments[exp_key] / run_name
                if (p / "metrics.json").exists():
                    m = load_metrics(p)
                    rows.append({"exp": exp_key, "trait": trait, "scalar": scalar,
                                 "run": run_name, **m})

    return pd.DataFrame(rows)


# ── sheet builders ─────────────────────────────────────────────────────────────

def build_sheet_overview(wb, df, title):
    ws = wb.create_sheet("Overview")
    ws.sheet_view.zoomScale = 110

    ws.merge_cells(f"A1:{get_column_letter(4 + len(METRICS))}1")
    c = ws["A1"]
    c.value     = f"{title} — Activation-Steered-Personas Experiment Overview"
    c.fill      = HEADER_FILL
    c.font      = Font(bold=True, color="FFFFFF", size=12)
    c.alignment = CENTER

    for ci, lbl in enumerate(["Experiment", "Trait", "Scalar", "# Runs"] + METRIC_LABELS, 1):
        style_header(ws.cell(row=2, column=ci, value=lbl))

    row_idx = 3
    for exp_name in ("baseline", "sa"):
        sub  = df[df.exp == exp_name]
        fill = EXP_FILLS[exp_name]
        for _, r in sub.iterrows():
            ws.cell(row=row_idx, column=1, value=exp_name).fill = fill
            ws.cell(row=row_idx, column=2, value="—").fill      = fill
            ws.cell(row=row_idx, column=3, value="—").fill      = fill
            ws.cell(row=row_idx, column=4, value=1).fill        = fill
            for ci, key in enumerate(METRIC_KEYS, 5):
                write_metric(ws, row_idx, ci, key, r[key], fill)
            row_idx += 1

    for exp_name in ("user-steer", "user-steer-sa"):
        sub  = df[df.exp == exp_name]
        fill = EXP_FILLS[exp_name]

        ws.merge_cells(f"A{row_idx}:D{row_idx}")
        c = ws.cell(row=row_idx, column=1,
                    value=f"{exp_name}  ▸  MEAN over all 40 runs")
        c.fill = HEADER2_FILL
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = LEFT
        c.border    = thin_border()
        for ci, key in enumerate(METRIC_KEYS, 5):
            write_metric(ws, row_idx, ci, key, sub[key].mean(), HEADER2_FILL)
        row_idx += 1

        for scalar in SCALARS:
            s_sub = sub[sub.scalar == scalar]
            sfill = SCALAR_FILLS[scalar]
            ws.cell(row=row_idx, column=1, value=exp_name).fill           = sfill
            ws.cell(row=row_idx, column=2, value="ALL TRAITS (mean)").fill = sfill
            ws.cell(row=row_idx, column=3, value=scalar).fill             = sfill
            ws.cell(row=row_idx, column=4, value=len(s_sub)).fill         = sfill
            for ci, key in enumerate(METRIC_KEYS, 5):
                write_metric(ws, row_idx, ci, key, s_sub[key].mean(), sfill)
            row_idx += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    for ci in range(5, 5 + len(METRIC_LABELS)):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    freeze_and_zoom(ws, "E3")


def build_sheet_exp1_vs_exp2(wb, df):
    ws = wb.create_sheet("Exp1 vs Exp2 (System SA effect)")

    ws.merge_cells(f"A1:{get_column_letter(5)}1")
    c = ws["A1"]
    c.value     = "Effect of System Persona-Flow  |  Baseline (Exp1) vs Persona-Flow (Exp2)"
    c.fill      = HEADER_FILL
    c.font      = Font(bold=True, color="FFFFFF", size=12)
    c.alignment = CENTER

    for ci, h in enumerate(["Metric", "Baseline", "Persona-Flow", "Δ (SA – Base)", "Δ%"], 1):
        style_header(ws.cell(row=2, column=ci, value=h))

    base = df[df.exp == "baseline"].iloc[0]
    sa   = df[df.exp == "sa"].iloc[0]

    for ri, (key, label, mode) in enumerate(METRICS, 3):
        bv  = base[key]
        pfv = sa[key]
        delta     = (pfv - bv) if (bv is not None and pfv is not None) else None
        delta_pct = (delta / bv) if (delta is not None and bv and bv != 0) else None

        ws.cell(row=ri, column=1, value=label).fill   = SUBHEAD_FILL
        ws.cell(row=ri, column=1).font   = LABEL_FONT
        ws.cell(row=ri, column=1).border = thin_border()
        write_metric(ws, ri, 2, key, bv,  EXP_FILLS["baseline"])
        write_metric(ws, ri, 3, key, pfv, EXP_FILLS["sa"])
        write_delta(ws, ri, 4, key, delta)
        # Δ% always as percentage of baseline
        if delta_pct is not None:
            c2 = ws.cell(row=ri, column=5, value=delta_pct)
            c2.font      = Font(size=10, bold=True,
                                color="375623" if delta_pct >= 0 else "C00000")
            c2.alignment = CENTER
            c2.border    = thin_border()
            c2.number_format = "+0.00%;-0.00%;0.00%"
        else:
            c2 = ws.cell(row=ri, column=5, value="")
            c2.alignment = CENTER
            c2.border    = thin_border()

    for ci, w in zip(range(1, 6), [22, 14, 14, 14, 10]):
        ws.column_dimensions[get_column_letter(ci)].width = w
    freeze_and_zoom(ws, "B3")


def build_sheet_all_runs(wb, df, exp_name, sheet_name):
    ws = wb.create_sheet(sheet_name)

    ws.merge_cells(f"A1:{get_column_letter(2 + len(METRICS))}1")
    c = ws["A1"]
    c.value     = f"{sheet_name}  —  10 traits × 4 scalars (40 runs).  s0 = no steering"
    c.fill      = HEADER_FILL
    c.font      = Font(bold=True, color="FFFFFF", size=12)
    c.alignment = CENTER

    for ci, lbl in enumerate(["Trait", "Scalar"] + METRIC_LABELS, 1):
        style_header(ws.cell(row=2, column=ci, value=lbl))

    sub = df[df.exp == exp_name]
    row_idx = 3
    data_first = row_idx

    for trait in TRAITS:
        for scalar in SCALARS:
            r = sub[(sub.trait == trait) & (sub.scalar == scalar)]
            if r.empty:
                continue
            r = r.iloc[0]
            sfill = SCALAR_FILLS[scalar]
            for ci_off, val in enumerate([trait, scalar]):
                c2 = ws.cell(row=row_idx, column=ci_off + 1, value=val)
                c2.fill = sfill; c2.font = NORMAL_FONT; c2.border = thin_border()
            for ci, key in enumerate(METRIC_KEYS, 3):
                write_metric(ws, row_idx, ci, key, r[key], sfill)
            row_idx += 1

    data_last = row_idx - 1
    for ci in range(3, 3 + len(METRIC_KEYS)):
        apply_color_scale(ws, data_first, data_last, ci)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 7
    for ci in range(3, 3 + len(METRIC_LABELS)):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    freeze_and_zoom(ws, "C3")


def build_sheet_delta(wb, df):
    ws = wb.create_sheet("Exp3 vs Exp4 (Δ adds System SA)")

    ws.merge_cells(f"A1:{get_column_letter(2 + len(METRICS))}1")
    c = ws["A1"]
    c.value     = ("Δ = user-steer+SA  minus  user-steer  "
                   "|  positive = SA improves steered-user baseline")
    c.fill      = HEADER_FILL
    c.font      = Font(bold=True, color="FFFFFF", size=12)
    c.alignment = CENTER

    for ci, lbl in enumerate(["Trait", "Scalar"] + METRIC_LABELS, 1):
        style_header(ws.cell(row=2, column=ci, value=lbl))

    us   = df[df.exp == "user-steer"]
    us_sa = df[df.exp == "user-steer-sa"]
    row_idx = 3

    for trait in TRAITS:
        for scalar in SCALARS:
            r3 = us[  (us.trait   == trait) & (us.scalar   == scalar)]
            r4 = us_sa[(us_sa.trait == trait) & (us_sa.scalar == scalar)]
            if r3.empty or r4.empty:
                continue
            r3, r4 = r3.iloc[0], r4.iloc[0]
            sfill = SCALAR_FILLS[scalar]
            for ci_off, val in enumerate([trait, scalar]):
                c2 = ws.cell(row=row_idx, column=ci_off + 1, value=val)
                c2.fill = sfill; c2.font = NORMAL_FONT; c2.border = thin_border()
            for ci, key in enumerate(METRIC_KEYS, 3):
                v3 = r3[key]; v4 = r4[key]
                delta = (v4 - v3) if (v3 is not None and v4 is not None) else None
                write_delta(ws, row_idx, ci, key, delta)
            row_idx += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 7
    for ci in range(3, 3 + len(METRIC_LABELS)):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    freeze_and_zoom(ws, "C3")


def build_sheet_scalar_pivot(wb, df, exp_name, sheet_name):
    ws = wb.create_sheet(sheet_name)

    ws.merge_cells(f"A1:{get_column_letter(1 + len(METRICS))}1")
    c = ws["A1"]
    c.value = sheet_name; c.fill = HEADER_FILL
    c.font  = Font(bold=True, color="FFFFFF", size=12); c.alignment = CENTER

    for ci, lbl in enumerate(["Scalar"] + METRIC_LABELS, 1):
        style_header(ws.cell(row=2, column=ci, value=lbl))

    sub = df[df.exp == exp_name]
    row_idx = 3
    for scalar in SCALARS:
        s_sub = sub[sub.scalar == scalar]
        sfill = SCALAR_FILLS[scalar]
        c2 = ws.cell(row=row_idx, column=1, value=scalar)
        c2.fill = sfill; c2.font = LABEL_FONT; c2.border = thin_border()
        for ci, key in enumerate(METRIC_KEYS, 2):
            write_metric(ws, row_idx, ci, key, s_sub[key].mean(), sfill)
        row_idx += 1

    ws.column_dimensions["A"].width = 8
    for ci in range(2, 2 + len(METRIC_LABELS)):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    freeze_and_zoom(ws, "B3")


def build_sheet_trait_pivot(wb, df, exp_name, sheet_name):
    ws = wb.create_sheet(sheet_name)

    ws.merge_cells(f"A1:{get_column_letter(2 + len(METRICS))}1")
    c = ws["A1"]
    c.value     = f"{sheet_name}  —  rows = trait, sub-rows = scalar"
    c.fill      = HEADER_FILL
    c.font      = Font(bold=True, color="FFFFFF", size=12)
    c.alignment = CENTER

    for ci, lbl in enumerate(["Trait", "Scalar"] + METRIC_LABELS, 1):
        style_header(ws.cell(row=2, column=ci, value=lbl))

    sub = df[df.exp == exp_name]
    row_idx = 3

    for trait in TRAITS:
        t_sub = sub[sub.trait == trait]
        for ci_off, val in enumerate([trait, "mean"]):
            c2 = ws.cell(row=row_idx, column=ci_off + 1, value=val)
            c2.fill = SUBHEAD_FILL; c2.font = LABEL_FONT; c2.border = thin_border()
        for ci, key in enumerate(METRIC_KEYS, 3):
            write_metric(ws, row_idx, ci, key, t_sub[key].mean(), SUBHEAD_FILL)
        row_idx += 1

        for scalar in SCALARS:
            r = t_sub[t_sub.scalar == scalar]
            if r.empty:
                continue
            r = r.iloc[0]
            sfill = SCALAR_FILLS[scalar]
            ws.cell(row=row_idx, column=1, value="").fill  = sfill
            ws.cell(row=row_idx, column=1).border = thin_border()
            c2 = ws.cell(row=row_idx, column=2, value=scalar)
            c2.fill = sfill; c2.font = NORMAL_FONT; c2.border = thin_border()
            for ci, key in enumerate(METRIC_KEYS, 3):
                write_metric(ws, row_idx, ci, key, r[key], sfill)
            row_idx += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 7
    for ci in range(3, 3 + len(METRIC_LABELS)):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    freeze_and_zoom(ws, "C3")


def build_sheet_notes(wb):
    ws = wb.create_sheet("Metric Glossary")
    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = "Metric Glossary"; c.fill = HEADER_FILL
    c.font  = Font(bold=True, color="FFFFFF", size=13); c.alignment = CENTER

    notes = [
        ("Method Acc",      "method_accuracy",        "pct",    "API method name predicted exactly"),
        ("Key Acc",         "key_accuracy",            "pct",    "Slot keys predicted correctly (precision)"),
        ("Key IoU",         "key_iou",                 "pct",    "Slot keys: intersection-over-union"),
        ("Value Acc",       "value_accuracy",          "pct",    "Slot values predicted correctly (precision)"),
        ("Value IoU",       "value_iou",               "pct",    "Slot values: IoU"),
        ("Full API Acc",    "full_api_accuracy",       "pct",    "Entire API call matches exactly"),
        ("Full API IoU",    "full_api_accuracy_iou",   "pct",    "Entire API call correct using IoU relaxation"),
        ("Dialog Success",  "successful_dialogs_rate", "pct",    "% dialogs where all required API calls succeeded"),
        ("Dialog Succ IoU", "successful_dialogs_iou_rate", "pct","Same but with IoU relaxation"),
        ("Inform Acc",      "inform_accuracy",         "pct",    "System correctly informs user of slot values"),
        ("Inform+Search",   "inform_accuracy_with_search_results", "pct", "Inform acc when search results available"),
        ("BLEU",            "bleu",                    "pct",    "BLEU score of system utterances vs gold"),
        ("Completion",      "dialog_completion_rate",  "pct",    "% dialogs that reached natural end state"),
        ("BERT F1",         "bert_score.avg_f1",       "pct",    "BERTScore F1 over all turns"),
        ("BERT User F1",    "bert_score_user.avg_f1",  "pct",    "BERTScore F1 on user-side turns"),
        ("BERT System F1",  "bert_score_system.avg_f1","pct",    "BERTScore F1 on system-side turns"),
        ("Csat",            "constraint_satisfaction_pct", "pct100", "LLM-judged constraint satisfaction (0–100)"),
        ("Inform (LLM)",    "inform_score",            "raw",    "LLM-judged informativeness score (1–5)"),
        ("Truthful",        "truthfulness_score",      "raw",    "LLM-judged truthfulness score (1–5)"),
        ("User Sat",        "user_satisfaction_score", "raw",    "LLM-judged user satisfaction score (1–5)"),
        ("PPL",             "perplexity.perplexity",   "ppl",    "Perplexity of generated utterances (lower=better)"),
    ]

    for ci, lbl in enumerate(["Short Label", "JSON key", "Scale", "Description"], 1):
        c2 = ws.cell(row=2, column=ci, value=lbl)
        c2.fill = HEADER2_FILL; c2.font = HEADER_FONT

    for ri, (short, key, scale, desc) in enumerate(notes, 3):
        ws.cell(row=ri, column=1, value=short).font  = LABEL_FONT
        ws.cell(row=ri, column=2, value=key).font    = Font(name="Courier New", size=9)
        ws.cell(row=ri, column=3, value=scale).font  = NORMAL_FONT
        ws.cell(row=ri, column=4, value=desc).font   = NORMAL_FONT
        ws.cell(row=ri, column=4).alignment          = LEFT

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 60

    scalar_start = len(notes) + 5
    ws.cell(row=scalar_start, column=1, value="Scalar Legend").font = LABEL_FONT
    for ri, (scalar, meaning, fill) in enumerate([
        ("s-1", "Anti-persona / negative direction",  SCALAR_FILLS["s-1"]),
        ("s0",  "No steering — identical to baseline", SCALAR_FILLS["s0"]),
        ("s1",  "Mild positive steering",              SCALAR_FILLS["s1"]),
        ("s2",  "Strong positive steering",            SCALAR_FILLS["s2"]),
    ], scalar_start + 1):
        c2 = ws.cell(row=ri, column=1, value=scalar)
        c2.fill = fill; c2.font = LABEL_FONT; c2.border = thin_border()
        ws.cell(row=ri, column=2, value=meaning).font = NORMAL_FONT


# ── main ───────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(description="Build Activation-Steered-Personas results Excel workbook")
    p.add_argument("--exp-root",      default=str(_DEFAULT_BASE))
    p.add_argument("--baseline",      default=_DEFAULT_EXPERIMENTS["baseline"])
    p.add_argument("--sa",  default=_DEFAULT_EXPERIMENTS["sa"])
    p.add_argument("--user-steer",    default=_DEFAULT_EXPERIMENTS["user-steer"])
    p.add_argument("--user-steer-sa", default=_DEFAULT_EXPERIMENTS["user-steer-sa"])
    p.add_argument("--output",        default=_DEFAULT_OUTPUT)
    p.add_argument("--title",         default="Activation-Steered-Personas")
    return p.parse_args()


def main():
    args = parse_args()
    root = Path(args.exp_root)

    experiments = {
        "baseline":      root / getattr(args, "baseline")     / "runs" / "baseline",
        "sa":  root / getattr(args, "sa") / "runs" / "sa",
        "user-steer":    root / getattr(args, "user_steer")   / "runs",
        "user-steer-sa": root / getattr(args, "user_steer_sa")/ "runs",
    }

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    print("Loading metrics …")
    df = load_all(experiments)
    print(f"  Loaded {len(df)} runs")

    from openpyxl import Workbook
    wb = Workbook()
    del wb["Sheet"]

    print("Building sheets …")
    build_sheet_overview(wb, df, args.title)
    build_sheet_exp1_vs_exp2(wb, df)
    build_sheet_all_runs(wb, df, "user-steer",    "Exp3 User-Steer (all runs)")
    build_sheet_all_runs(wb, df, "user-steer-sa", "Exp4 User-Steer+SA (all runs)")
    build_sheet_delta(wb, df)
    build_sheet_scalar_pivot(wb, df, "user-steer",    "Exp3 by Scalar")
    build_sheet_scalar_pivot(wb, df, "user-steer-sa", "Exp4 by Scalar")
    build_sheet_trait_pivot(wb, df, "user-steer",    "Exp3 by Trait")
    build_sheet_trait_pivot(wb, df, "user-steer-sa", "Exp4 by Trait")
    build_sheet_notes(wb)

    wb.save(out_path)
    print(f"Saved → {out_path}")
    return df


if __name__ == "__main__":
    df = main()
